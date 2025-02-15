import streamlit as st
import qrcode
import pandas as pd
from io import BytesIO
import zipfile
import os
import base64
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

def generate_qr_code(url, product_name):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(url)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    return img

def create_download_link(data, filename, mime_type):
    b64 = base64.b64encode(data).decode()
    href = f'<a href="data:{mime_type};base64,{b64}" download="{filename}">{filename}をダウンロード</a>'
    return href

def save_to_excel(df, qr_images):
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "商品QRコード"

    # ヘッダーを設定
    ws['A1'] = 'ProductName'
    ws['B1'] = 'ProductUrl'
    ws['C1'] = 'QR Code'

    # 列幅を調整
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 40

    # データを書き込む
    for idx, (_, row) in enumerate(df.iterrows(), start=2):
        ws.cell(row=idx, column=1, value=row['ProductName'])
        ws.cell(row=idx, column=2, value=row['ProductUrl'])
        
        # QRコードを一時ファイルとして保存
        img_buffer = BytesIO()
        qr_images[idx-2].save(img_buffer, format='PNG')
        img_buffer.seek(0)
        
        # QRコードを画像としてExcelに挿入
        img = XLImage(img_buffer)
        img.width = 150  # 画像サイズを調整
        img.height = 150
        ws.add_image(img, f'C{idx}')
        
        # 行の高さを調整
        ws.row_dimensions[idx].height = 120

    # Excelファイルをバイトストリームとして保存
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

def main():
    st.title("商品QRコード生成")
    
    # CSVファイルのアップロード
    uploaded_file = st.file_uploader("CSVファイルをアップロードしてください", type=['csv'])
    
    if uploaded_file is not None:
        try:
            # Try different encodings
            encodings = ['utf-8', 'shift-jis', 'cp932']
            df = None
            
            for encoding in encodings:
                try:
                    df = pd.read_csv(uploaded_file, encoding=encoding)
                    break
                except UnicodeDecodeError:
                    continue
                finally:
                    uploaded_file.seek(0)
            
            if df is None:
                st.error("CSVファイルのエンコーディングを認識できませんでした。UTF-8またはShift-JISで保存してください。")
                return
            
            if len(df.columns) < 2:
                st.error("CSVファイルにはProductNameとProductUrlの2列が必要です")
                return
                
            # 最初の2列を使用
            df = df.iloc[:, :2]
            df.columns = ['ProductName', 'ProductUrl']
            
            # QRコードの生成と表示
            st.write("### 生成されたQRコード")
            
            # ZIP作成用のメモリストリーム
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
                # QRコードのオブジェクトを格納するリスト
                qr_images = []
                
                # 各商品のQRコード生成
                for idx, row in df.iterrows():
                    product_name = row['ProductName']
                    url = row['ProductUrl']
                    
                    # QRコード生成
                    qr_img = generate_qr_code(url, product_name)
                    qr_images.append(qr_img)
                    
                    # 画像をバイトストリームに変換してZIPに追加
                    img_buffer = BytesIO()
                    qr_img.save(img_buffer, format='PNG')
                    
                    # ZIPに追加
                    filename = f"{product_name}_qr.png"
                    zip_file.writestr(filename, img_buffer.getvalue())
                    
                    # Streamlitに表示
                    col1, col2 = st.columns([1, 2])
                    with col1:
                        st.image(img_buffer, caption=product_name, width=200)
                    with col2:
                        st.write(f"**商品名:** {product_name}")
                        st.write(f"**URL:** [{url}]({url})")
                    st.divider()
                
                # Excelファイルの作成
                excel_buffer = save_to_excel(df, qr_images)
            
            # ダウンロードリンクの作成
            zip_buffer.seek(0)
            st.markdown("### ダウンロード")
            col1, col2 = st.columns(2)
            with col1:
                st.markdown(create_download_link(zip_buffer.getvalue(), "product_qrcodes.zip", "application/zip"), unsafe_allow_html=True)
            with col2:
                st.markdown(create_download_link(excel_buffer.getvalue(), "products_with_qr.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"), unsafe_allow_html=True)
            
        except Exception as e:
            st.error(f"エラーが発生しました: {str(e)}")
            st.write("CSVファイルの形式を確認してください。以下の形式が必要です：")
            st.code("ProductName,ProductUrl\nProduct A,https://example.com/productA\nProduct B,https://example.com/productB")

if __name__ == "__main__":
    st.set_page_config(page_title="商品QRコード生成", layout="wide")
    main()
